import os
import json
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
import requests
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill

# Get the directory this script is in
SCRIPT_DIR = Path(__file__).parent.absolute()

# Load environment variables from the same directory as the script
load_dotenv(SCRIPT_DIR / '.env')

class JobDataProcessor:
    def __init__(self):
        self.api_key = os.getenv('API_KEY')
        self.bearer_token = os.getenv('BEARER_TOKEN')
        self.api_endpoint = os.getenv('API_ENDPOINT')
        
        # Create output directories if they don't exist
        self.output_dir = SCRIPT_DIR / 'output'
        self.json_dir = self.output_dir / 'json'
        self.excel_dir = self.output_dir / 'excel'
        
        for dir_path in [self.output_dir, self.json_dir, self.excel_dir]:
            dir_path.mkdir(parents=True, exist_ok=True)

    def fetch_data(self):
        """Fetch data from the API endpoint"""
        headers = {
            'Authorization': f'Bearer {self.bearer_token}',
            'Apikey': self.api_key
        }
        
        try:
            response = requests.get(self.api_endpoint, headers=headers)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Error fetching data: {e}")
            return None

    def format_timestamp(self, timestamp_str):
        """Convert timestamp string to date format"""
        try:
            # Parse the timestamp string to datetime
            dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            # Format as YYYY-MM-DD
            return dt.strftime('%Y-%m-%d')
        except (ValueError, AttributeError):
            return timestamp_str

    def save_data(self, data):
        """Save data to both JSON and Excel formats with customizations"""
        if not data:
            return False
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save as JSON
        json_path = self.json_dir / f'job_data_{timestamp}.json'
        with open(json_path, 'w') as f:
            json.dump(data, f, indent=4)
        
        # Create DataFrame and process data
        df = pd.DataFrame(data)
        
        # Format created_at timestamp
        if 'created_at' in df.columns:
            df['created_at'] = df['created_at'].apply(self.format_timestamp)
        
        # Remove specified columns
        columns_to_remove = ['id', 'is_our_job', 'user_id']
        df = df.drop(columns=[col for col in columns_to_remove if col in df.columns])
        
        # Add applied_at column
        df['applied_at'] = ''
        
        # Save as Excel
        excel_path = self.excel_dir / f'job_data_{timestamp}.xlsx'
        df.to_excel(excel_path, index=False)
        
        # Add data validation and formatting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Define status options
        status_options = ['rejected', 'offer', 'contacted', 'applied', 'interviewing']
        status_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(status_options)}"',
            allow_blank=True
        )
        
        # Add date validation for applied_at column
        date_validation = DataValidation(
            type="date",
            allow_blank=True,
            error="Please enter a valid date (YYYY-MM-DD)"
        )
        
        # Find columns
        status_col = None
        applied_at_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value and cell.value.lower() == 'status':
                status_col = idx
            elif cell.value and cell.value.lower() == 'applied_at':
                applied_at_col = idx
        
        # Add validations
        if status_col:
            status_validation.add(f'{chr(64 + status_col)}2:{chr(64 + status_col)}{ws.max_row}')
            ws.add_data_validation(status_validation)
        
        if applied_at_col:
            date_validation.add(f'{chr(64 + applied_at_col)}2:{chr(64 + applied_at_col)}{ws.max_row}')
            ws.add_data_validation(date_validation)
            
            # Add light gray background to date column to indicate it's editable
            light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            for row in range(2, ws.max_row + 1):
                ws[f'{chr(64 + applied_at_col)}{row}'].fill = light_gray_fill
        
        # Save the modified Excel file
        wb.save(excel_path)
        
        print(f"Data saved successfully:")
        print(f"JSON: {json_path}")
        print(f"Excel: {excel_path}")
        return True

    def run(self):
        """Main execution method"""
        print("Fetching job data...")
        data = self.fetch_data()
        if data:
            self.save_data(data)
        else:
            print("No data to process")

if __name__ == "__main__":
    processor = JobDataProcessor()
    processor.run() 